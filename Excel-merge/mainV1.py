import os
import sys
import site
from copy import copy

def setup_tkdnd():
    # 获取 site-packages 目录
    site_packages = site.getsitepackages()
    for site_pkg in site_packages:
        # 修改为使用 win-x86 目录
        tkdnd_path = os.path.join(site_pkg, 'tkinterdnd2', 'tkdnd', 'win-x86')
        if os.path.exists(tkdnd_path):
            # 将 tkdnd 目录添加到系统路径
            os.environ['PATH'] = tkdnd_path + os.pathsep + os.environ.get('PATH', '')
            return True
    return False

# 在导入 tkinterdnd2 之前设置环境
setup_tkdnd()
import tkinterdnd2 as tkdnd
import pandas as pd
from openpyxl import load_workbook
import tkinter as tk
from tkinter import messagebox, filedialog
from tkinter import ttk
import time
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

class DragDropGUI(tkdnd.Tk):
    def __init__(self):
        super().__init__()
        
        self.title("Excel文件合并工具")
        self.geometry("800x600")  # 增加窗口尺寸
        self.configure(padx=30, pady=30)  # 增加内边距
        
        # 设置字体 - 使用更美观的字体设置
        self.title_font = ("微软雅黑", 24, "bold")     # 增大标题字体
        self.hint_font = ("微软雅黑", 12)              # 增大提示文字
        self.normal_font = ("微软雅黑", 11)            # 增大普通文本
        
        self.template_path = None
        self.input_dir = None
        self.output_path = None
        self.create_widgets()
        # 设置默认提示文本
        self.save_path_entry.insert(0, "请输入或选择保存位置")
        
    def create_widgets(self):
        # 标题和说明
        title_frame = ttk.Frame(self)
        title_frame.pack(fill=tk.X, pady=(0, 15))
        
        title_label = ttk.Label(title_frame, text="Excel文件合并工具", font=self.title_font)
        title_label.pack()
        
        hint_label = ttk.Label(title_frame, 
                              text="请按顺序选择：1.模板文件 2.输入文件夹 3.保存位置",
                              font=self.hint_font)
        hint_label.pack(pady=(10, 0))
        
        # 创建拖拽区域的主框架
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 左侧拖拽区域
        left_frame = ttk.LabelFrame(main_frame, text="模板文件", padding=10)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        # 添加文件名和删除按钮的框架
        template_name_frame = ttk.Frame(left_frame)
        template_name_frame.pack(fill=tk.X)
        
        self.template_name_label = ttk.Label(template_name_frame, text="", anchor="w", font=self.normal_font)
        self.template_name_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        self.template_clear_button = ttk.Label(template_name_frame, text="×", cursor="hand2", font=self.normal_font)
        self.template_clear_button.pack(side=tk.RIGHT, padx=(5, 0))
        self.template_clear_button.bind('<Button-1>', lambda e: self.clear_template())
        self.template_clear_button.configure(foreground='red')  # 设置×为红色
        self.template_clear_button.pack_forget()  # 初始时隐藏
        
        self.template_label = ttk.Label(left_frame, 
                                      text="请拖入模板文件\n或点击选择",
                                      background='#F0F0F0',
                                      padding=(10, 30),  # 替换 height，使用 padding 来控制高度
                                      justify='center',
                                      font=self.normal_font)
        self.template_label.pack(fill=tk.BOTH, expand=True)
        self.template_label.drop_target_register(tkdnd.DND_FILES)
        self.template_label.dnd_bind('<<Drop>>', self.drop_template)
        self.template_label.bind('<Button-1>', lambda e: self.select_template())
        
        # 右侧拖拽区域
        right_frame = ttk.LabelFrame(main_frame, text="输入文件/文件夹", padding=10)
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))
        
        # 添加文件名和删除按钮的框架
        input_name_frame = ttk.Frame(right_frame)
        input_name_frame.pack(fill=tk.X)
        
        self.input_name_label = ttk.Label(input_name_frame, text="", anchor="w", font=self.normal_font)
        self.input_name_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        self.input_clear_button = ttk.Label(input_name_frame, text="×", cursor="hand2", font=self.normal_font)
        self.input_clear_button.pack(side=tk.RIGHT, padx=(5, 0))
        self.input_clear_button.bind('<Button-1>', lambda e: self.clear_input())
        self.input_clear_button.configure(foreground='red')  # 设置×为红色
        self.input_clear_button.pack_forget()  # 初始时隐藏
        
        self.input_label = ttk.Label(right_frame,
                                   text="请拖入Excel文件或文件夹\n或点击选择",
                                   background='#F0F0F0',
                                   padding=(10, 30),  # 替换 height，使用 padding 来控制高度
                                   justify='center',
                                   font=self.normal_font)
        self.input_label.pack(fill=tk.BOTH, expand=True)
        self.input_label.drop_target_register(tkdnd.DND_FILES)
        self.input_label.dnd_bind('<<Drop>>', self.drop_input)
        self.input_label.bind('<Button-1>', lambda e: self.select_input_dir())
        
        # 底部区域使用Grid布局
        bottom_frame = ttk.Frame(self)
        bottom_frame.pack(fill=tk.X, pady=(10, 0))
        bottom_frame.grid_columnconfigure(0, weight=1)
        
        # 保存路径显示和选择按钮
        save_frame = ttk.LabelFrame(bottom_frame, text="保存位置", padding=5)
        save_frame.pack(fill=tk.X, pady=(0, 10))
        save_frame.grid_columnconfigure(0, weight=1)
        
        # 将Label改为Entry
        self.save_path_entry = ttk.Entry(save_frame)
        self.save_path_entry.grid(row=0, column=0, sticky='ew', padx=5)
        self.save_path_entry.bind('<Return>', lambda e: self.validate_save_path())  # 按回车时验证路径
        self.save_path_entry.bind('<FocusOut>', lambda e: self.validate_save_path())  # 失去焦点时验证路径
        
        self.save_button = ttk.Button(
            save_frame,
            text="选择位置",
            command=self.select_save_location,
            width=10
        )
        self.save_button.grid(row=0, column=1, padx=5)
        
        # 状态显示和合并按钮
        status_frame = ttk.Frame(bottom_frame)
        status_frame.pack(fill=tk.X)
        status_frame.grid_columnconfigure(0, weight=1)
        
        self.status_label = ttk.Label(status_frame, text="请选择模板文件", foreground="red")
        self.status_label.grid(row=0, column=0, sticky='w', padx=5)
        
        self.merge_button = ttk.Button(
            status_frame,
            text="开始合并",
            command=self.start_merge,
            width=20
        )
        self.merge_button.grid(row=0, column=1, padx=5)
    
    def drop_template(self, event):
        file_path = event.data
        file_path = file_path.strip('{}')
        if file_path.lower().endswith('.xlsx'):
            self.template_path = file_path
            self.template_name_label.configure(text=f"已选择: {os.path.basename(self.template_path)}")
            self.template_label.configure(text="")
            self.template_clear_button.pack(side=tk.RIGHT, padx=(5, 0))
            
            # 设置默认保存路径
            default_output = get_default_output_path(self.template_path)
            self.save_path_entry.delete(0, tk.END)
            self.save_path_entry.insert(0, default_output)
            self.output_path = default_output
            
            self.update_status()
        else:
            messagebox.showwarning("警告", "请选择Excel文件(.xlsx)作为模板")
    
    def drop_input(self, event):
        paths = event.data.split('}') if '}' in event.data else [event.data]
        paths = [p.strip('{ ') for p in paths if p.strip('{ ')]  # 清理路径字符串
        
        valid_paths = []
        for path in paths:
            if os.path.isfile(path) and path.lower().endswith('.xlsx'):
                valid_paths.append(path)
            elif os.path.isdir(path):
                valid_paths.append(path)
        
        if valid_paths:
            # 如果已经有选择的文件，则追加新文件
            if self.input_dir:
                self.input_dir.extend(valid_paths)
            else:
                self.input_dir = valid_paths
            
            # 更新显示文本
            total_files = 0
            if len(self.input_dir) == 1 and os.path.isdir(self.input_dir[0]):
                folder_files = [f for f in os.listdir(self.input_dir[0]) if f.lower().endswith('.xlsx')]
                total_files = len(folder_files)
                if folder_files:
                    self.input_name_label.configure(text=f"已选择文件夹，包含 {total_files} 个Excel文件")
                else:
                    self.input_name_label.configure(text="已选择文件夹 (文件夹为空)")
            else:
                total_files = len([p for p in self.input_dir if os.path.isfile(p)])
                folder_files = sum(len([f for f in os.listdir(p) if f.lower().endswith('.xlsx')]) 
                                 for p in self.input_dir if os.path.isdir(p))
                total_files += folder_files
                self.input_name_label.configure(text=f"已选择 {total_files} 个Excel文件")
            
            self.input_label.configure(text="")
            self.input_clear_button.pack(side=tk.RIGHT, padx=(5, 0))  # 显示删除按钮
            self.update_status()
        else:
            messagebox.showwarning("警告", "请选择Excel文件(.xlsx)或文件夹")
    
    def select_template(self):
        path = filedialog.askopenfilename(
            title="选择模板文件",
            filetypes=[("Excel文件", "*.xlsx")]
        )
        if path:
            self.template_path = path
            self.template_name_label.configure(text=f"已选择: {os.path.basename(path)}")
            self.template_label.configure(text="")
            self.template_clear_button.pack(side=tk.RIGHT, padx=(5, 0))
            
            # 设置默认保存路径
            default_output = get_default_output_path(self.template_path)
            self.save_path_entry.delete(0, tk.END)
            self.save_path_entry.insert(0, default_output)
            self.output_path = default_output
            
            self.update_status()
    
    def select_input_dir(self):
        # 先尝试选择文件
        paths = filedialog.askopenfilenames(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx")]
        )
        
        if paths:  # 如果选择了文件
            self.input_dir = list(paths)
            self.update_input_text(f"已选择 {len(paths)} 个Excel文件")
            self.input_label.configure(text="")
            self.update_status()
        else:  # 如果没选文件，尝试选择文件夹
            path = filedialog.askdirectory(title="选择文件夹")
            if path:
                self.input_dir = [path]
                # 获取文件夹中的Excel文件数量
                folder_files = [f for f in os.listdir(path) if f.lower().endswith('.xlsx')]
                if folder_files:
                    self.update_input_text(f"已选择文件夹，包含 {len(folder_files)} 个Excel文件")
                else:
                    self.update_input_text("已选择文件夹 (文件夹为空)")
                self.input_label.configure(text="")
                self.update_status()
    
    def select_save_location(self):
        output_path = filedialog.asksaveasfilename(
            title="选择保存位置",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")]
        )
        if output_path:
            self.output_path = output_path
            # 更新Entry中的文本
            self.save_path_entry.delete(0, tk.END)
            self.save_path_entry.insert(0, output_path)
            self.update_status()
    
    def update_status(self):
        if not self.template_path:
            self.status_label.configure(text="请选择模板文件", foreground="red")
            self.merge_button.state(['disabled'])
        elif not self.input_dir:
            self.status_label.configure(text="请选择输入文件夹", foreground="red")
            self.merge_button.state(['disabled'])
        elif not self.output_path:
            self.status_label.configure(text="请选择保存位置", foreground="red")
            self.merge_button.state(['disabled'])
        else:
            self.status_label.configure(text="可以开始合并了！", foreground="green")
            self.merge_button.state(['!disabled'])
    
    def start_merge(self):
        if not (self.template_path and self.input_dir):
            return
            
        if not self.output_path:
            self.select_save_location()
            if not self.output_path:
                return
            
        merge_excel_files(self.template_path, self.input_dir, self.output_path)

    def update_template_text(self, text):
        self.template_name_label.configure(text=text)

    def update_input_text(self, text):
        """更新输入文件显示文本"""
        self.input_name_label.configure(text=text)

    def validate_save_path(self):
        """验证手动输入的保存路径"""
        path = self.save_path_entry.get().strip()
        if not path:
            self.output_path = None
            self.update_status()
            return False
            
        # 确保路径以.xlsx结尾
        if not path.lower().endswith('.xlsx'):
            path += '.xlsx'
            self.save_path_entry.delete(0, tk.END)
            self.save_path_entry.insert(0, path)
        
        try:
            # 检查目录是否存在
            directory = os.path.dirname(path)
            if directory and not os.path.exists(directory):
                messagebox.showwarning("警告", "保存路径所在文件夹不存在！")
                return False
                
            self.output_path = path
            self.update_status()
            return True
        except Exception as e:
            messagebox.showwarning("警告", f"无效的保存路径: {str(e)}")
            return False

    def clear_template(self):
        """清除已选择的模板文件"""
        self.template_path = None
        self.template_name_label.configure(text="")
        self.template_label.configure(text="请拖入模板文件\n或点击选择")
        self.template_clear_button.pack_forget()  # 隐藏删除按钮
        self.update_status()

    def clear_input(self):
        """清除已选择的输入文件"""
        self.input_dir = None
        self.input_name_label.configure(text="")
        self.input_label.configure(text="请拖入Excel文件或文件夹\n或点击选择")
        self.input_clear_button.pack_forget()  # 隐藏删除按钮
        self.update_status()

def validate_template(template_path, input_path):
    """验证输入文件是否符合模板格式"""
    try:
        template = pd.read_excel(template_path, nrows=1)
        input_file = pd.read_excel(input_path, nrows=1)
        return list(template.columns) == list(input_file.columns)
    except Exception as e:
        messagebox.showerror("错误", f"验证文件 {input_path} 时出错: {str(e)}")
        return False

def get_default_output_path(template_path):
    """根据模板路径生成默认的输出路径"""
    base, ext = os.path.splitext(template_path)
    return f"{base}_merge{ext}"

def merge_excel_files(template_path, input_paths, output_path=None):
    try:
        # 如果没有指定输出路径，使用默认路径
        if output_path is None:
            output_path = get_default_output_path(template_path)
        
        print(f"\n开始合并文件:")
        print(f"模板文件: {template_path}")
        print(f"输入路径: {input_paths}")
        print(f"输出路径: {output_path}")
        
        # 读取模板数据和样式
        template_df = pd.read_excel(template_path, sheet_name=0)
        print(f"\n模板文件第一个sheet的列: {list(template_df.columns)}")
        
        # 获取模板中的样式信息
        template_wb = load_workbook(template_path)
        template_ws = template_wb.worksheets[0]
        
        # 获取模板中每列的样式（包括标题行和数据行）
        template_header_styles = []  # 存储标题行样式
        template_data_styles = []    # 存储数据行样式
        
        # 获取每列的标题样式（第一行）
        for col in range(1, len(template_df.columns) + 1):
            header_cell = template_ws.cell(row=1, column=col)
            template_header_styles.append(header_cell)
        
        # 获取模板最后一行的样式作为标准格式
        last_row = 1
        for row in template_ws.iter_rows():
            if any(cell.value is not None for cell in row):
                last_row = row[0].row
        
        # 获取每列最后一行的样式
        for col in range(1, len(template_df.columns) + 1):
            cell = template_ws.cell(row=last_row, column=col)
            # 如果最后一行的单元格为空，向上查找最近的非空单元格
            if cell.value is None:
                for row in range(last_row-1, 1, -1):
                    temp_cell = template_ws.cell(row=row, column=col)
                    if temp_cell.value is not None:
                        cell = temp_cell
                        break
            template_data_styles.append(cell)
        
        # 创建模板行的集合用于比较
        template_rows = set()
        for row in template_df.iloc[1:].values:
            row_values = []
            for val in row:
                if pd.isna(val):
                    row_values.append('')
                elif isinstance(val, (int, float)):
                    row_values.append(str(val))
                else:
                    row_values.append(str(val) if val is not None else '')
            template_rows.add(tuple(row_values))
        
        print(f"模板文件数据行数: {len(template_rows)}")
        
        # 复制模板文件到输出位置
        import shutil
        shutil.copy2(template_path, output_path)
        
        # 读取输出文件
        result_wb = load_workbook(output_path)
        result_ws = result_wb.worksheets[0]
        
        # 找到最后一个非空行
        last_row = 1
        for row in result_ws.iter_rows():
            if any(cell.value is not None for cell in row):
                last_row = row[0].row
        
        current_row = last_row + 1
        print(f"开始写入的行号: {current_row}")
        
        # 处理输入文件
        input_files = []
        for path in input_paths:
            if os.path.isfile(path):
                input_files.append(path)
            else:
                folder_files = [os.path.join(path, f) 
                              for f in os.listdir(path) 
                              if f.endswith('.xlsx')]
                input_files.extend(folder_files)
        
        print(f"\n找到的输入文件: {len(input_files)}")
        success_count = 0
        failed_files = []
        file_row_ranges = {}
        
        # 处理每个输入文件
        for file_path in input_files:
            try:
                print(f"\n处理文件: {file_path}")
                file_name = os.path.basename(file_path)
                
                # 读取数据
                df = pd.read_excel(file_path, sheet_name=0)
                print(f"读取到 {len(df)} 行数据")
                
                # 验证列是否匹配
                if list(df.columns) != list(template_df.columns):
                    error_msg = "列不匹配"
                    print(f"错误: {error_msg}")
                    failed_files.append((file_name, error_msg))
                    continue
                
                # 处理数据行
                data_rows = []
                seen_rows = set()
                
                for row in df.iloc[1:].values:
                    row_values = []
                    for val in row:
                        if pd.isna(val):
                            row_values.append('')
                        elif isinstance(val, (int, float)):
                            row_values.append(str(val))
                        else:
                            row_values.append(str(val) if val is not None else '')
                    
                    row_tuple = tuple(row_values)
                    if row_tuple not in template_rows and row_tuple not in seen_rows:
                        data_rows.append(row)
                        seen_rows.add(row_tuple)
                
                if not data_rows:
                    print("警告: 文件中没有新的数据行")
                    continue
                
                print(f"写入 {len(data_rows)} 行数据")
                
                # 记录起始行
                start_row = current_row
                
                # 写入数据
                for row_data in data_rows:
                    for col, value in enumerate(row_data, 1):
                        cell = result_ws.cell(row=current_row, column=col)
                        cell.value = value if pd.notna(value) else None
                    
                    # 添加来源文件信息
                    source_cell = result_ws.cell(row=current_row, column=len(template_df.columns) + 1)
                    source_cell.value = file_name
                    
                    current_row += 1
                
                # 记录行范围
                if data_rows:
                    file_row_ranges[file_name] = (start_row, current_row - 1)
                
                success_count += 1
                print(f"文件处理成功")
                
            except Exception as e:
                error_msg = f"{str(e)}\n位置: {traceback.extract_tb(sys.exc_info()[2])[-1]}"
                print(f"处理文件时出错: {error_msg}")
                failed_files.append((file_name, error_msg))
        
        # 如果有成功处理的文件，应用样式
        if success_count > 0:
            # 添加来源文件列标题并应用样式
            source_col = len(template_df.columns) + 1
            source_header = result_ws.cell(row=1, column=source_col)
            source_header.value = "来源文件"
            
            # 应用标题行样式
            for col, header_style in enumerate(template_header_styles, 1):
                if header_style:
                    header_cell = result_ws.cell(row=1, column=col)
                    # 确保完整复制字体属性
                    if header_style.font:
                        new_font = copy(header_style.font)
                        header_cell.font = new_font
                    header_cell.alignment = copy(header_style.alignment)
                    header_cell.fill = copy(header_style.fill)
                    header_cell.border = copy(header_style.border)
                    if hasattr(header_style, 'number_format'):
                        header_cell.number_format = header_style.number_format
            
            # 应用数据行样式
            template_row_count = len(template_df)  # 获取模板的行数
            
            for row in range(2, current_row):
                # 跳过模板中原有的行，保持其原有格式
                if row <= template_row_count + 1:  # +1 是因为包含标题行
                    continue
                    
                # 对新增的行应用样式
                for col, template_cell in enumerate(template_data_styles):
                    if template_cell:
                        cell = result_ws.cell(row=row, column=col + 1)
                        # 确保完整复制字体属性
                        if template_cell.font:
                            new_font = copy(template_cell.font)
                            cell.font = new_font
                        cell.alignment = copy(template_cell.alignment)
                        cell.fill = copy(template_cell.fill)
                        cell.border = copy(template_cell.border)
                        if hasattr(template_cell, 'number_format'):
                            cell.number_format = template_cell.number_format
                
                # 为来源文件列应用第一列的数据样式
                source_cell = result_ws.cell(row=row, column=source_col)
                if template_data_styles[0]:
                    if template_data_styles[0].font:
                        new_font = copy(template_data_styles[0].font)
                        source_cell.font = new_font
                    source_cell.alignment = copy(template_data_styles[0].alignment)
                    source_cell.fill = copy(template_data_styles[0].fill)
                    source_cell.border = copy(template_data_styles[0].border)
                    if hasattr(template_data_styles[0], 'number_format'):
                        source_cell.number_format = template_data_styles[0].number_format
            
            # 调整来源文件列宽度
            source_col_letter = get_column_letter(source_col)
            result_ws.column_dimensions[source_col_letter].width = 15
            
            # 合并相同来源文件的单元格
            for file_name, (start_row, end_row) in file_row_ranges.items():
                if start_row < end_row:
                    merge_range = f"{source_col_letter}{start_row}:{source_col_letter}{end_row}"
                    result_ws.merge_cells(merge_range)
                    merged_cell = result_ws.cell(row=start_row, column=source_col)
                    merged_cell.alignment = Alignment(vertical='center')
            
            # 保存结果
            result_wb.save(output_path)
            
            # 根据是否有失败文件显示不同的成功消息
            if failed_files:
                # 保存失败记录
                timestamp = time.strftime("%Y%m%d_%H%M%S")
                output_dir = os.path.dirname(output_path)
                fail_log_path = os.path.join(output_dir, f'合并失败记录_{timestamp}.txt')
                
                with open(fail_log_path, 'w', encoding='utf-8') as f:
                    f.write(f"合并失败的文件列表 (共 {len(failed_files)} 个)：\n")
                    f.write("="*50 + "\n")
                    for i, (file_name, reason) in enumerate(failed_files, 1):
                        f.write(f"{i}. {file_name}: {reason}\n")
                    f.write(f"\n记录时间：{time.strftime('%Y-%m-%d %H:%M:%S')}")
                
                messagebox.showinfo("部分成功", 
                    f"已成功合并 {success_count} 个文件到 {output_path}\n"
                    f"有 {len(failed_files)} 个文件合并失败\n"
                    f"失败记录已保存到：{fail_log_path}")
            else:
                messagebox.showinfo("成功", f"已成功合并 {success_count} 个文件到 {output_path}")
        else:
            messagebox.showwarning("警告", "没有成功合并任何文件！")
            
    except Exception as e:
        import traceback
        error_msg = f"{str(e)}\n位置: {traceback.extract_tb(sys.exc_info()[2])[-1]}"
        print(f"合并文件时出错: {error_msg}")
        messagebox.showerror("错误", f"合并文件时出错: {error_msg}")

def main():
    app = DragDropGUI()
    app.mainloop()

if __name__ == "__main__":
    main()
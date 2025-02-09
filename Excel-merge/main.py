import os
import site
from copy import copy
import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def setup_tkdnd():
    for site_pkg in site.getsitepackages():
        tkdnd_path = os.path.join(site_pkg, 'tkinterdnd2', 'tkdnd', 'win-x86')
        if os.path.exists(tkdnd_path):
            os.environ['PATH'] = tkdnd_path + os.pathsep + os.environ.get('PATH', '')
            return True
    return False

setup_tkdnd()
import tkinterdnd2 as tkdnd

class DragDropGUI(tkdnd.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel文件合并工具")
        self.geometry("800x600")
        self.configure(padx=30, pady=30)
        
        self.title_font = ("微软雅黑", 24, "bold")
        self.hint_font = ("微软雅黑", 12)
        self.normal_font = ("微软雅黑", 11)
        
        self.template_path = None
        self.input_dir = None
        self.output_path = None
        self.create_widgets()
        self.save_path_entry.insert(0, "请输入或选择保存位置")

    def create_widgets(self):
        # 标题和说明
        title_frame = ttk.Frame(self)
        title_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(title_frame, text="Excel文件合并工具", font=self.title_font).pack()
        ttk.Label(title_frame, text="请按顺序选择：1.模板文件 2.输入文件夹 3.保存位置", 
                 font=self.hint_font).pack(pady=(10, 0))
        
        # 创建拖拽区域
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 左侧模板区域
        left_frame = ttk.LabelFrame(main_frame, text="模板文件", padding=10)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        template_name_frame = ttk.Frame(left_frame)
        template_name_frame.pack(fill=tk.X)
        
        self.template_name_label = ttk.Label(template_name_frame, text="", anchor="w", 
                                           font=self.normal_font)
        self.template_name_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        self.template_clear_button = ttk.Label(template_name_frame, text="×", cursor="hand2",
                                             font=self.normal_font, foreground='red')
        self.template_clear_button.bind('<Button-1>', lambda e: self.clear_template())
        self.template_clear_button.pack_forget()
        
        self.template_label = ttk.Label(left_frame, text="请拖入模板文件\n或点击选择",
                                      background='#F0F0F0', padding=(10, 30),
                                      justify='center', font=self.normal_font)
        self.template_label.pack(fill=tk.BOTH, expand=True)
        self.template_label.drop_target_register(tkdnd.DND_FILES)
        self.template_label.dnd_bind('<<Drop>>', self.drop_template)
        self.template_label.bind('<Button-1>', lambda e: self.select_template())
        
        # 右侧输入区域
        right_frame = ttk.LabelFrame(main_frame, text="输入文件/文件夹", padding=10)
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))
        
        input_name_frame = ttk.Frame(right_frame)
        input_name_frame.pack(fill=tk.X)
        
        self.input_name_label = ttk.Label(input_name_frame, text="", anchor="w",
                                        font=self.normal_font)
        self.input_name_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        self.input_clear_button = ttk.Label(input_name_frame, text="×", cursor="hand2",
                                          font=self.normal_font, foreground='red')
        self.input_clear_button.bind('<Button-1>', lambda e: self.clear_input())
        self.input_clear_button.pack_forget()
        
        self.input_label = ttk.Label(right_frame, text="请拖入Excel文件或文件夹\n或点击选择",
                                   background='#F0F0F0', padding=(10, 30),
                                   justify='center', font=self.normal_font)
        self.input_label.pack(fill=tk.BOTH, expand=True)
        self.input_label.drop_target_register(tkdnd.DND_FILES)
        self.input_label.dnd_bind('<<Drop>>', self.drop_input)
        self.input_label.bind('<Button-1>', lambda e: self.select_input_dir())
        
        # 底部区域
        bottom_frame = ttk.Frame(self)
        bottom_frame.pack(fill=tk.X, pady=(10, 0))
        
        save_frame = ttk.LabelFrame(bottom_frame, text="保存位置", padding=5)
        save_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.save_path_entry = ttk.Entry(save_frame)
        self.save_path_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        self.save_path_entry.bind('<Return>', lambda e: self.validate_save_path())
        self.save_path_entry.bind('<FocusOut>', lambda e: self.validate_save_path())
        
        ttk.Button(save_frame, text="选择位置", command=self.select_save_location,
                  width=10).pack(side=tk.RIGHT, padx=5)
        
        status_frame = ttk.Frame(bottom_frame)
        status_frame.pack(fill=tk.X)
        
        self.status_label = ttk.Label(status_frame, text="请选择模板文件", foreground="red")
        self.status_label.pack(side=tk.LEFT, padx=5)
        
        self.merge_button = ttk.Button(status_frame, text="开始合并",
                                     command=self.start_merge, width=20)
        self.merge_button.pack(side=tk.RIGHT, padx=5)
        self.merge_button.state(['disabled'])

    def drop_template(self, event):
        file_path = event.data.strip('{}')
        if file_path.lower().endswith('.xlsx'):
            self.template_path = file_path
            self.template_name_label.configure(text=f"已选择: {os.path.basename(file_path)}")
            self.template_label.configure(text="")
            self.template_clear_button.pack(side=tk.RIGHT, padx=(5, 0))
            
            default_output = os.path.splitext(file_path)[0] + "_merge.xlsx"
            self.save_path_entry.delete(0, tk.END)
            self.save_path_entry.insert(0, default_output)
            self.output_path = default_output
            
            self.update_status()
        else:
            messagebox.showwarning("警告", "请选择Excel文件(.xlsx)作为模板")

    def drop_input(self, event):
        paths = event.data.split('}') if '}' in event.data else [event.data]
        paths = [p.strip('{ ') for p in paths if p.strip('{ ')]
        
        valid_paths = []
        for path in paths:
            if os.path.isfile(path) and path.lower().endswith('.xlsx'):
                valid_paths.append(path)
            elif os.path.isdir(path):
                valid_paths.append(path)
        
        if valid_paths:
            if self.input_dir:
                self.input_dir.extend(valid_paths)
            else:
                self.input_dir = valid_paths
            
            total_files = len([p for p in self.input_dir if os.path.isfile(p)])
            folder_files = sum(len([f for f in os.listdir(p) if f.lower().endswith('.xlsx')]) 
                             for p in self.input_dir if os.path.isdir(p))
            total_files += folder_files
            
            self.input_name_label.configure(text=f"已选择 {total_files} 个Excel文件")
            self.input_label.configure(text="")
            self.input_clear_button.pack(side=tk.RIGHT, padx=(5, 0))
            self.update_status()
        else:
            messagebox.showwarning("警告", "请选择Excel文件(.xlsx)或文件夹")

    def select_template(self):
        path = filedialog.askopenfilename(title="选择模板文件",
                                        filetypes=[("Excel文件", "*.xlsx")])
        if path:
            self.template_path = path
            self.template_name_label.configure(text=f"已选择: {os.path.basename(path)}")
            self.template_label.configure(text="")
            self.template_clear_button.pack(side=tk.RIGHT, padx=(5, 0))
            
            default_output = os.path.splitext(path)[0] + "_merge.xlsx"
            self.save_path_entry.delete(0, tk.END)
            self.save_path_entry.insert(0, default_output)
            self.output_path = default_output
            
            self.update_status()

    def select_input_dir(self):
        paths = filedialog.askopenfilenames(title="选择Excel文件",
                                          filetypes=[("Excel文件", "*.xlsx")])
        if paths:
            self.input_dir = list(paths)
            self.input_name_label.configure(text=f"已选择 {len(paths)} 个Excel文件")
            self.input_label.configure(text="")
            self.input_clear_button.pack(side=tk.RIGHT, padx=(5, 0))
            self.update_status()
        else:
            path = filedialog.askdirectory(title="选择文件夹")
            if path:
                self.input_dir = [path]
                folder_files = [f for f in os.listdir(path) if f.lower().endswith('.xlsx')]
                self.input_name_label.configure(
                    text=f"已选择文件夹，包含 {len(folder_files)} 个Excel文件")
                self.input_label.configure(text="")
                self.input_clear_button.pack(side=tk.RIGHT, padx=(5, 0))
                self.update_status()

    def select_save_location(self):
        output_path = filedialog.asksaveasfilename(
            title="选择保存位置",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")]
        )
        if output_path:
            self.output_path = output_path
            self.save_path_entry.delete(0, tk.END)
            self.save_path_entry.insert(0, output_path)
            self.update_status()

    def update_status(self):
        if not self.template_path:
            self.status_label.configure(text="请选择模板文件", foreground="red")
            self.merge_button.state(['disabled'])
        elif not self.input_dir:
            self.status_label.configure(text="请选择输入文件夹", foreground="red")
            self.merge_button.state(['disabled'])
        elif not self.output_path:
            self.status_label.configure(text="请选择保存位置", foreground="red")
            self.merge_button.state(['disabled'])
        else:
            self.status_label.configure(text="可以开始合并了！", foreground="green")
            self.merge_button.state(['!disabled'])

    def validate_save_path(self):
        path = self.save_path_entry.get().strip()
        if not path:
            self.output_path = None
            self.update_status()
            return False
            
        if not path.lower().endswith('.xlsx'):
            path += '.xlsx'
            self.save_path_entry.delete(0, tk.END)
            self.save_path_entry.insert(0, path)
        
        try:
            directory = os.path.dirname(path)
            if directory and not os.path.exists(directory):
                messagebox.showwarning("警告", "保存路径所在文件夹不存在！")
                return False
                
            self.output_path = path
            self.update_status()
            return True
        except Exception as e:
            messagebox.showwarning("警告", f"无效的保存路径: {str(e)}")
            return False

    def clear_template(self):
        self.template_path = None
        self.template_name_label.configure(text="")
        self.template_label.configure(text="请拖入模板文件\n或点击选择")
        self.template_clear_button.pack_forget()
        self.update_status()

    def clear_input(self):
        self.input_dir = None
        self.input_name_label.configure(text="")
        self.input_label.configure(text="请拖入Excel文件或文件夹\n或点击选择")
        self.input_clear_button.pack_forget()
        self.update_status()

    def start_merge(self):
        if not (self.template_path and self.input_dir):
            return
            
        if not self.output_path:
            self.select_save_location()
            if not self.output_path:
                return
            
        merge_excel_files(self.template_path, self.input_dir, self.output_path)

def merge_excel_files(template_path, input_paths, output_path):
    try:
        template_df = pd.read_excel(template_path)
        template_wb = load_workbook(template_path)
        template_ws = template_wb.worksheets[0]
        
        template_header_styles = []
        template_data_styles = []
        
        for col in range(1, len(template_df.columns) + 1):
            template_header_styles.append(template_ws.cell(row=1, column=col))
        
        last_row = max((row[0].row for row in template_ws.iter_rows() 
                       if any(cell.value is not None for cell in row)), default=1)
        
        for col in range(1, len(template_df.columns) + 1):
            cell = template_ws.cell(row=last_row, column=col)
            if cell.value is None:
                for row in range(last_row-1, 1, -1):
                    temp_cell = template_ws.cell(row=row, column=col)
                    if temp_cell.value is not None:
                        cell = temp_cell
                        break
            template_data_styles.append(cell)
        
        template_rows = {tuple(str(val) if val is not None else '' 
                             for val in row) for row in template_df.iloc[1:].values}
        
        import shutil
        shutil.copy2(template_path, output_path)
        
        result_wb = load_workbook(output_path)
        result_ws = result_wb.worksheets[0]
        
        last_row = max((row[0].row for row in result_ws.iter_rows() 
                       if any(cell.value is not None for cell in row)), default=1)
        
        current_row = last_row + 1
        input_files = []
        
        for path in input_paths:
            if os.path.isfile(path):
                input_files.append(path)
            else:
                input_files.extend(os.path.join(path, f) for f in os.listdir(path) 
                                 if f.endswith('.xlsx'))
        
        success_count = 0
        failed_files = []
        file_row_ranges = {}
        
        for file_path in input_files:
            try:
                df = pd.read_excel(file_path)
                if list(df.columns) != list(template_df.columns):
                    failed_files.append((os.path.basename(file_path), "列不匹配"))
                    continue
                
                data_rows = []
                seen_rows = set()
                
                for row in df.iloc[1:].values:
                    row_tuple = tuple(str(val) if val is not None else '' for val in row)
                    if row_tuple not in template_rows and row_tuple not in seen_rows:
                        data_rows.append(row)
                        seen_rows.add(row_tuple)
                
                if not data_rows:
                    continue
                
                start_row = current_row
                
                for row_data in data_rows:
                    for col, value in enumerate(row_data, 1):
                        cell = result_ws.cell(row=current_row, column=col)
                        cell.value = value if pd.notna(value) else None
                    
                    source_cell = result_ws.cell(row=current_row, 
                                               column=len(template_df.columns) + 1)
                    source_cell.value = os.path.basename(file_path)
                    
                    current_row += 1
                
                if data_rows:
                    file_row_ranges[os.path.basename(file_path)] = (start_row, current_row - 1)
                
                success_count += 1
                
            except Exception as e:
                failed_files.append((os.path.basename(file_path), str(e)))
        
        if success_count > 0:
            source_col = len(template_df.columns) + 1
            source_header = result_ws.cell(row=1, column=source_col)
            source_header.value = "来源文件"
            
            for col, header_style in enumerate(template_header_styles, 1):
                if header_style:
                    header_cell = result_ws.cell(row=1, column=col)
                    if header_style.font:
                        header_cell.font = copy(header_style.font)
                    header_cell.alignment = copy(header_style.alignment)
                    header_cell.fill = copy(header_style.fill)
                    header_cell.border = copy(header_style.border)
                    if hasattr(header_style, 'number_format'):
                        header_cell.number_format = header_style.number_format
            
            template_row_count = len(template_df)
            
            for row in range(2, current_row):
                if row <= template_row_count + 1:
                    continue
                    
                for col, template_cell in enumerate(template_data_styles):
                    if template_cell:
                        cell = result_ws.cell(row=row, column=col + 1)
                        if template_cell.font:
                            cell.font = copy(template_cell.font)
                        cell.alignment = copy(template_cell.alignment)
                        cell.fill = copy(template_cell.fill)
                        cell.border = copy(template_cell.border)
                        if hasattr(template_cell, 'number_format'):
                            cell.number_format = template_cell.number_format
                
                source_cell = result_ws.cell(row=row, column=source_col)
                if template_data_styles[0]:
                    if template_data_styles[0].font:
                        source_cell.font = copy(template_data_styles[0].font)
                    source_cell.alignment = copy(template_data_styles[0].alignment)
                    source_cell.fill = copy(template_data_styles[0].fill)
                    source_cell.border = copy(template_data_styles[0].border)
                    if hasattr(template_data_styles[0], 'number_format'):
                        source_cell.number_format = template_data_styles[0].number_format
            
            source_col_letter = get_column_letter(source_col)
            result_ws.column_dimensions[source_col_letter].width = 15
            
            for file_name, (start_row, end_row) in file_row_ranges.items():
                if start_row < end_row:
                    merge_range = f"{source_col_letter}{start_row}:{source_col_letter}{end_row}"
                    result_ws.merge_cells(merge_range)
                    merged_cell = result_ws.cell(row=start_row, column=source_col)
                    merged_cell.alignment = Alignment(vertical='center')
            
            result_wb.save(output_path)
            
            # 保存失败记录
            if failed_files:
                error_log_path = os.path.splitext(output_path)[0] + "_错误记录.txt"
                with open(error_log_path, 'w', encoding='utf-8') as f:
                    f.write(f"合并失败的文件记录：\n")
                    f.write(f"时间：{pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write("-" * 50 + "\n")
                    for file_name, error in failed_files:
                        f.write(f"文件：{file_name}\n")
                        f.write(f"原因：{error}\n")
                        f.write("-" * 50 + "\n")
                
                messagebox.showinfo("部分成功", 
                    f"已成功合并 {success_count} 个文件\n"
                    f"有 {len(failed_files)} 个文件合并失败\n"
                    f"失败记录已保存至：{os.path.basename(error_log_path)}")
            else:
                messagebox.showinfo("成功", f"已成功合并 {success_count} 个文件")
        else:
            messagebox.showwarning("警告", "没有成功合并任何文件！")
            
    except Exception as e:
        messagebox.showerror("错误", f"合并文件时出错: {str(e)}")

def main():
    app = DragDropGUI()
    app.mainloop()

if __name__ == "__main__":
    main()